import csv
import zipfile
from pypdf import PdfReader
from script_os import ZIP_RESOURCES
from openpyxl import load_workbook


def test_read_pdf():
    with zipfile.ZipFile(ZIP_RESOURCES) as zip_file: #открытие архива
        with zip_file.open('test-pdf.pdf') as file: #цикл на открытие файла в архиве
            reader = PdfReader(file)
            text_pdf = reader.pages[0].extract_text()
            assert "TESTING PDF FILE FOR DOWNLOAD" in text_pdf



def test_read_xlsx():
    with zipfile.ZipFile(ZIP_RESOURCES) as zip_file:
        with zip_file.open('file.xlsx') as file:
            workbook = load_workbook(file)
            sheet = workbook.active
            assert sheet.cell(row=16, column=2).value == 'Kina'



def test_read_csv():
    with zipfile.ZipFile(ZIP_RESOURCES) as zip_file:
        with zip_file.open('users.csv') as file:

            content = file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]

            assert 'Oleg' in second_row[0]


